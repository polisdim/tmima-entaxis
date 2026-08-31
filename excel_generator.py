# -*- coding: utf-8 -*-
"""
Excel (.xlsx) Central Session Log Generator for Inclusion Class (Τμήμα Ένταξης)
School: ΔΗΜ.Ω.Σ. Γυμνάσιο Ξάνθης
Teacher: Δημήτριος Πολυχρόνης (ΠΕ03.ΕΑΕ)
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_sessions_excel(observations, students, output_path):
    wb = openpyxl.Workbook()
    
    # 1. Main Sheet: Ημερολόγιο Συνεδριών Τ.Ε.
    ws1 = wb.active
    ws1.title = "Συνεδρίες Τ.Ε."
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Header Block
    ws1.merge_cells('A1:L1')
    title_cell = ws1['A1']
    title_cell.value = "ΗΜΕΡΟΛΟΓΙΟ ΣΥΝΕΔΡΙΩΝ & ΔΙΔΑΚΤΙΚΩΝ ΠΑΡΕΜΒΑΣΕΩΝ ΤΜΗΜΑΤΟΣ ΕΝΤΑΞΗΣ (2026-2027)"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30
    
    ws1.merge_cells('A2:L2')
    sub_cell = ws1['A2']
    sub_cell.value = f"Σχολική Μονάδα: ΔΗΜ.Ω.Σ. Γυμνάσιο Ξάνθης | Εκπαιδευτικός: Δημήτριος Πολυχρόνης (ΠΕ03.ΕΑΕ) | Εξαγωγή: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub_cell.font = Font(name='Calibri', size=10, italic=True, color='D9E1F2')
    sub_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 20

    # Column Headers
    headers = [
        ("Α/Α", 6),
        ("Ημερομηνία & Ώρα", 18),
        ("Τύπος", 14),
        ("Μαθητής / Τμήμα", 22),
        ("Τάξη/Τμήμα", 12),
        ("Θεματικός Τομέας", 18),
        ("Επίπεδο Bloom", 16),
        ("Αναπαράσταση (Duval)", 20),
        ("Εμπόδιο (Brousseau)", 24),
        ("Στρατηγική / Τεχνική", 22),
        ("Αποτέλεσμα", 14),
        ("Περιγραφή & Σημειώσεις Εκπαιδευτικού", 45)
    ]
    
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='1F4E79')
    thin_border = Border(
        left=Side(style='thin', color='BDD7EE'),
        right=Side(style='thin', color='BDD7EE'),
        top=Side(style='thin', color='BDD7EE'),
        bottom=Side(style='medium', color='1F4E79')
    )
    data_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    ws1.row_dimensions[3].height = 26
    for col_idx, (h_text, width) in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = width

    # Fill Data
    student_map = {s['id']: s for s in students}
    sorted_obs = sorted(observations, key=lambda x: x.get('timestamp', ''), reverse=True)
    
    pos_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # soft green
    pos_font = Font(name='Calibri', size=11, bold=True, color='375623')
    part_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # soft yellow
    part_font = Font(name='Calibri', size=11, bold=True, color='7F6000')
    neg_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # soft red
    neg_font = Font(name='Calibri', size=11, bold=True, color='C65911')

    for row_idx, o in enumerate(sorted_obs, 4):
        ws1.row_dimensions[row_idx].height = 22
        st_id = o.get('student_id')
        st_info = student_map.get(st_id, {})
        
        target_type = o.get('target_type', 'individual')
        if target_type == 'individual':
            target_type_str = 'Ατομική'
        elif target_type == 'group':
            target_type_str = 'Ομάδα Τ.Ε.'
        elif target_type == 'class':
            target_type_str = 'Τμήμα Τ.Ε.'
        else:
            target_type_str = str(target_type)

        raw_ts = o.get('timestamp', '')
        dt_str = raw_ts[:16].replace('T', ' ') if raw_ts else '-'
        
        st_name = o.get('student_name', st_info.get('name', '-'))
        class_sec = o.get('class_section') or st_info.get('class_section', st_info.get('grade', '-'))
        domain = o.get('domain_name', o.get('domain_id', '-'))
        bloom = o.get('bloom_name', o.get('bloom_id', '-'))
        duval = o.get('duval_name', o.get('duval_register', '-'))
        obst = o.get('obstacle', '-')
        strat = o.get('strategy_name', '-')
        if o.get('technique'):
            strat += f" ({o.get('technique')})"
        
        out_code = o.get('outcome_code', '+')
        out_label = f"[{out_code}] Θετικό" if out_code == '+' else (f"[{out_code}] Μερικό" if out_code == '~' else f"[{out_code}] Δυσκολία")
        notes = o.get('raw_text', '')

        row_values = [
            row_idx - 3,
            dt_str,
            target_type_str,
            st_name,
            class_sec,
            domain,
            bloom,
            duval,
            obst,
            strat,
            out_label,
            notes
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = data_border
            cell.font = Font(name='Calibri', size=10)
            
            if col_idx in [1, 3, 5, 11]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
            if col_idx == 11:
                if out_code == '+':
                    cell.fill = pos_fill
                    cell.font = pos_font
                elif out_code == '~':
                    cell.fill = part_fill
                    cell.font = part_font
                elif out_code == '-':
                    cell.fill = neg_fill
                    cell.font = neg_font

    if sorted_obs:
        ws1.auto_filter.ref = f"A3:L{len(sorted_obs) + 3}"

    # 2. Second Sheet: Πρόοδος Στόχων Ε.Π.Ε. & Ρουμπρίκες 4 Επιπέδων
    ws2 = wb.create_sheet(title="Στόχοι Ε.Π.Ε. & Ρουμπρίκα")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells('A1:H1')
    t2 = ws2['A1']
    t2.value = "ΕΞΑΤΟΜΙΚΕΥΜΕΝΟ ΠΡΟΓΡΑΜΜΑ ΕΚΠΑΙΔΕΥΣΗΣ (Ε.Π.Ε.) - ΡΟΥΜΠΡΙΚΑ 4 ΕΠΙΠΕΔΩΝ"
    t2.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    t2.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28
    
    iep_headers = [
        ("Α/Α", 6),
        ("Μαθητής/τρια", 20),
        ("Τάξη/Τμήμα", 12),
        ("Θεματικός Τομέας", 18),
        ("Διδακτικός Στόχος Ε.Π.Ε.", 40),
        ("Στάθμη Ρουμπρίκας", 22),
        ("Επίπεδο (1-4)", 14),
        ("Κατάσταση", 16)
    ]
    ws2.row_dimensions[2].height = 24
    for col_idx, (h_text, width) in enumerate(iep_headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = width

    rubric_names = {
        1: "1. Αρχικό (Πλήρης καθοδήγηση)",
        2: "2. Αναδυόμενο (Μερική βοήθεια)",
        3: "3. Ικανό (Αυτόνομη επίλυση)",
        4: "4. Γενικευμένο (Πλήρης κατάκτηση)"
    }

    iep_row = 3
    for st in students:
        st_full = f"{st.get('name','')} {st.get('surname','')}".strip()
        st_sec = st.get('class_section', st.get('grade','-'))
        targets = st.get('iep_targets', [])
        for t in targets:
            ws2.row_dimensions[iep_row].height = 22
            r_level = t.get('rubric_level', 2)
            r_name = rubric_names.get(r_level, rubric_names[2])
            
            row_data = [
                iep_row - 2,
                st_full,
                st_sec,
                t.get('area', 'Μαθηματικά'),
                t.get('target', '-'),
                r_name,
                f"Επίπεδο {r_level}/4",
                t.get('status', 'Σε εξέλιξη')
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=iep_row, column=col_idx, value=val)
                cell.border = data_border
                cell.font = Font(name='Calibri', size=10)
                if col_idx in [1, 3, 7, 8]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            iep_row += 1

    if iep_row > 3:
        ws2.auto_filter.ref = f"A2:H{iep_row - 1}"

    # Save to path
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path
